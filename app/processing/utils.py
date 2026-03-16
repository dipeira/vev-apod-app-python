"""
Processing pipeline:
  Step 1  excel_to_pdf  — LibreOffice headless converts Excel → PDF (landscape, exact formatting)
  Step 2  clean_pdf     — removes blank pages (≤3 words), returns kept-page indices
  Step 3  create_index  — scans clean PDF for AFM + AMKA → CSV index
"""
import csv
import logging
import os
import re
import shutil
import signal
import subprocess
import sys
import tempfile
import threading
import time
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo

def _now_athens():
    return datetime.now(ZoneInfo('Europe/Athens')).replace(tzinfo=None)

import PyPDF2

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# In-memory job state  { year_data_id: {progress, detail, abort} }
# ---------------------------------------------------------------------------
_state: dict = {}
_state_lock = threading.Lock()

_threads: dict = {}          # year_data_id → Thread
_threads_lock = threading.Lock()

# Global processing slot — only one pipeline may run at a time
_processing_lock = threading.Lock()


def claim_processing_slot() -> bool:
    """Try to claim the global processing slot. Returns True if successful."""
    return _processing_lock.acquire(blocking=False)


def release_processing_slot():
    """Release the global processing slot (called from the pipeline thread)."""
    try:
        _processing_lock.release()
    except RuntimeError:
        pass  # already released


def is_any_processing() -> bool:
    """Return True if a pipeline is currently running."""
    acquired = _processing_lock.acquire(blocking=False)
    if acquired:
        _processing_lock.release()
        return False
    return True


def _set(yd_id, **kw):
    if yd_id is None:
        return
    with _state_lock:
        _state.setdefault(yd_id, {}).update(kw)


def get_state(yd_id: int) -> dict:
    with _state_lock:
        return dict(_state.get(yd_id, {}))


def request_abort(yd_id: int):
    _set(yd_id, abort=True)


def wait_for_abort(yd_id: int, timeout: float = 15):
    """Signal abort and wait up to `timeout` seconds for the thread to stop."""
    request_abort(yd_id)
    with _threads_lock:
        t = _threads.get(yd_id)
    if t and t.is_alive():
        t.join(timeout=timeout)


def _aborted(yd_id) -> bool:
    if yd_id is None:
        return False
    with _state_lock:
        return _state.get(yd_id, {}).get('abort', False)


class _Abort(Exception):
    pass


# ---------------------------------------------------------------------------
# Cross-platform process termination
# ---------------------------------------------------------------------------
def _kill_proc(proc):
    """Terminate a subprocess and all its children (works on Linux and Windows)."""
    try:
        if sys.platform != 'win32':
            os.killpg(os.getpgid(proc.pid), signal.SIGTERM)
        else:
            proc.terminate()
    except Exception:
        proc.terminate()
    try:
        proc.wait(timeout=5)
    except Exception:
        proc.kill()


# ---------------------------------------------------------------------------
# LibreOffice discovery
# ---------------------------------------------------------------------------
_LO_CANDIDATES = [
    # Linux / Docker
    'libreoffice', 'soffice',
    '/usr/bin/libreoffice', '/usr/bin/soffice',
    '/usr/local/bin/libreoffice',
    # Windows
    r'C:\Program Files\LibreOffice\program\soffice.exe',
    r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    # macOS
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',
]


def _find_libreoffice():
    for c in _LO_CANDIDATES:
        if shutil.which(c) or os.path.isfile(c):
            return c
    return None


# ---------------------------------------------------------------------------
# Helper: optimize XLSX for PDF generation (Landscape + Autofit)
# ---------------------------------------------------------------------------
def _preprocess_xlsx(src_path: str, dst_path: str):
    """
    Patch XLSX to prevent ### on Linux/Docker while maintaining readable text size.
    Uses openpyxl to:
      1. Convert numbers/dates to strings so LibreOffice never renders ###.
      2. Set number_format to '@' (Text).
      3. Set shrink_to_fit on all cells.
      4. Set Landscape + FitToWidth=1, FitToHeight=0 (auto).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    import copy
    from datetime import date, datetime

    # Load with data_only=True to get formula results
    try:
        wb = openpyxl.load_workbook(src_path, data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(src_path)

    for ws in wb.worksheets:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = automatic height

        col_widths = {}  # {column_index: max_length}

        # Iterate all cells
        for row in ws.iter_rows():
            for col_idx, cell in enumerate(row):
                # 1. Force Shrink to Fit + Disable Wrap Text
                # This is the key: it allows LibreOffice to shrink font per-cell
                # instead of requiring a massive column width that shrinks the whole page.
                if cell.alignment:
                    new_align = copy.copy(cell.alignment)
                    new_align.shrink_to_fit = True
                    new_align.wrap_text = False
                    cell.alignment = new_align
                else:
                    cell.alignment = Alignment(shrink_to_fit=True, wrap_text=False)

                # 2. Convert Numbers/Dates to Strings to prevent '###'
                val = cell.value
                if val is not None:
                    # If it's a number, force it to string so it never renders as ###
                    if isinstance(val, (int, float)):
                        # Format logic: integers stay integers, floats get 2 decimals
                        if isinstance(val, float):
                            cell.value = f"{val:.2f}"
                        else:
                            cell.value = str(val)
                        cell.data_type = 's'  # Force String type
                        cell.number_format = '@'  # Force Text formatting
                    elif isinstance(val, (datetime, date)):
                        cell.value = val.strftime('%d/%m/%Y')
                        cell.data_type = 's'
                        cell.number_format = '@'

                    # 3. Measure content for column width
                    val_str = str(cell.value) if cell.value is not None else ''
                    length = len(val_str)
                    if length > col_widths.get(col_idx, 0):
                        col_widths[col_idx] = length

        # Apply widths
        for col_idx, max_len in col_widths.items():
            column_letter = get_column_letter(col_idx + 1)
            # Multiplier 1.8 prevents ### on Linux fonts without making text too small via FitToPage
            ws.column_dimensions[column_letter].width = (max_len + 5) * 1.8

    wb.save(dst_path)


# ---------------------------------------------------------------------------
# Step 1: Excel → PDF via LibreOffice (preserves exact sheet formatting)
# ---------------------------------------------------------------------------
def excel_to_pdf(excel_path, pdf_path, yd_id=None):
    """
    Convert Excel to PDF using LibreOffice.
    Pre-processes the file to enforce Landscape and widen columns (fixing ####).
    Handles .xls by converting to .xlsx first.
    Returns (success: bool, message: str, page_count: int)
    """
    lo = _find_libreoffice()
    if not lo:
        return (False,
                'LibreOffice δεν βρέθηκε. Εγκαταστήστε το LibreOffice και ξαναδοκιμάστε.',
                0)

    _set(yd_id, progress=6,
         detail='Βήμα 1/3: Προετοιμασία αρχείου (Landscape & Columns)…')

    out_dir = os.path.dirname(pdf_path)
    os.makedirs(out_dir, exist_ok=True)

    ext      = os.path.splitext(excel_path)[1].lower()
    lo_input = None
    temp_files = []
    lo_profile_dir = None

    try:
        if sys.platform == 'win32':
            # Windows: pass file directly to LibreOffice.
            # LibreOffice reads XLS/XLSX print settings (landscape, fit-to-page) natively.
            # Any XLSX intermediate step would lose those settings.
            lo_input = excel_path

        else:
            # Linux/Docker: need XLSX so we can patch column widths and page setup.
            src_xlsx = excel_path
            if ext == '.xls':
                _set(yd_id, detail='Βήμα 1/3: Μετατροπή .xls σε .xlsx…')
                try:
                    src_xlsx = _convert_xls_to_xlsx(excel_path)
                    temp_files.append(src_xlsx)
                except Exception as e:
                    return False, f'Αποτυχία μετατροπής .xls: {e}', 0

            optimized_xlsx = tempfile.mktemp(suffix='.xlsx', prefix='optimized_')
            temp_files.append(optimized_xlsx)
            try:
                _preprocess_xlsx(src_xlsx, optimized_xlsx)
                lo_input = optimized_xlsx
                logger.info('Optimized XLSX created: %s', optimized_xlsx)
            except Exception as e:
                logger.warning('Preprocessing failed (%s), using src', e)
                lo_input = src_xlsx

        _set(yd_id, detail='Βήμα 1/3: Δημιουργία PDF με LibreOffice…')

        # Private temp profile prevents concurrent-run conflicts in LibreOffice
        lo_profile_dir = tempfile.mkdtemp(prefix='lo_profile_')
        
        # Build a correct file:// URI for both Linux (/tmp/...) and Windows (C:\...)
        lo_profile_posix = lo_profile_dir.replace(os.sep, '/')
        if not lo_profile_posix.startswith('/'):
            lo_profile_posix = '/' + lo_profile_posix   # Windows: /C:/...
        lo_profile_uri = f'file://{lo_profile_posix}'   # file:///tmp/... or file:///C:/...

        # Kill any stale soffice processes left from a previous timed-out run.
        # On Linux only — on Windows this would be too aggressive.
        if sys.platform != 'win32':
            subprocess.run(['pkill', '-9', '-f', 'soffice'], capture_output=True)
            time.sleep(0.5)  # give the OS a moment to reap them

        proc = subprocess.Popen(
            [
                lo, '--headless',
                '--norestore',           # don't try to recover stale sessions
                '--nofirststartwizard',  # skip setup wizard on fresh profile dirs
                f'-env:UserInstallation={lo_profile_uri}',
                '--convert-to', 'pdf',
                '--outdir', out_dir,
                lo_input,
            ],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            start_new_session=True,   # creates a process group on Linux
            # On Linux/Docker only: give LO a clean home dir so it doesn't touch
            # the real user home. On Windows LO uses APPDATA/USERPROFILE, not HOME.
            env={**os.environ, 'HOME': lo_profile_dir} if sys.platform != 'win32' else None,
        )

        start = time.time()
        last_update = start
        while proc.poll() is None:
            if _aborted(yd_id):
                _kill_proc(proc)
                raise _Abort()
            now = time.time()
            if now - start > 1200:
                _kill_proc(proc)
                return False, 'LibreOffice timeout (>20 λεπτά).', 0
            if now - last_update >= 5:
                elapsed = int(now - start)
                _set(yd_id, detail=f'Βήμα 1/3: LibreOffice εκτελείται… ({elapsed}s)')
                last_update = now
            time.sleep(0.3)

        stdout_data, stderr_data = proc.communicate()
        if proc.returncode != 0:
            err = (stderr_data or stdout_data or 'Άγνωστο σφάλμα').strip()
            return False, f'LibreOffice error: {err}', 0

        base   = os.path.splitext(os.path.basename(lo_input))[0]
        lo_out = os.path.join(out_dir, f'{base}.pdf')

        if not os.path.exists(lo_out):
            return False, 'LibreOffice δεν παρήγαγε αρχείο PDF.', 0

        if lo_out != pdf_path:
            shutil.move(lo_out, pdf_path)

        with open(pdf_path, 'rb') as f:
            count = len(PyPDF2.PdfReader(f).pages)

        _set(yd_id, progress=33,
             detail=f'✓ Βήμα 1/3: PDF δημιουργήθηκε ({count} σελίδες)')
        return True, f'{count} σελίδες δημιουργήθηκαν.', count

    except _Abort:
        raise
    except subprocess.TimeoutExpired:
        return False, 'LibreOffice timeout (>20 λεπτά).', 0
    except Exception as e:
        logger.exception('excel_to_pdf failed')
        return False, str(e), 0
    finally:
        if lo_profile_dir:
            shutil.rmtree(lo_profile_dir, ignore_errors=True)
        for tmp in temp_files:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except OSError: pass

# ---------------------------------------------------------------------------
# Step 2: Clean PDF — remove blank pages (≤3 words)
# ---------------------------------------------------------------------------
def clean_pdf(input_path, output_path, yd_id=None):
    """
    Returns (success, msg, kept_count).
    Same logic as cleanPdf-dias.py.
    """
    try:
        with open(input_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            total   = len(reader.pages)
            writer  = PyPDF2.PdfWriter()
            kept = removed = 0

            for i, page in enumerate(reader.pages):
                if _aborted(yd_id):
                    raise _Abort()
                text  = re.sub(r'\s\s+', ' ',
                               (page.extract_text() or '').replace('\n', ' ').strip())
                words = text.split()
                if len(words) > 3:
                    writer.add_page(page)
                    kept += 1
                else:
                    removed += 1
                if i % 20 == 0 or i == total - 1:
                    pct  = 33 + int((i + 1) / total * 33)
                    flag = '✓' if len(words) > 3 else '✗ κενή'
                    _set(yd_id, progress=pct,
                         detail=(f'Βήμα 2/3: Σελίδα {i+1}/{total} [{flag}]  '
                                 f'Κρατήθηκαν: {kept}  Αφαιρέθηκαν: {removed}'))

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'wb') as f:
            writer.write(f)
        return True, f'{kept} σελίδες (αφαιρέθηκαν {removed} κενές).', kept

    except _Abort:
        raise
    except Exception as e:
        logger.exception('clean_pdf failed')
        return False, str(e), 0


# ---------------------------------------------------------------------------
# Step 3: Create CSV index
# ---------------------------------------------------------------------------
# Exact same patterns as createIndexFromPdf-dias.py
_AFM_PAT  = re.compile(r'\b\d{9}\b')   # 9-digit AFM


def create_index(pdf_path, csv_path, yd_id=None):
    """
    Scan each PDF page using the same logic as createIndexFromPdf-dias.py:
      - Iterate tokens; keep overwriting afm with each 9-digit match.
      - Record first 11-digit match as amka.
      - Break when both found.
    Result: afm = last 9-digit token found before amka (= employee personal AFM,
    since org AFM appears first and gets overwritten by employee AFM).

    Writes semicolon-delimited CSV: afm;amka;page
    Returns (success, msg, matched_count)
    """
    try:
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            total  = len(reader.pages)
            rows   = [['afm', 'amka', 'page']]
            count  = 0

            for i, page in enumerate(reader.pages):
                if _aborted(yd_id):
                    raise _Abort()

                tokens = re.sub(
                    r'\s\s+', ' ',
                    (page.extract_text() or '').replace('\n', ' ').strip()
                ).split()

                afm = amka = ''
                afm_found = amka_found = False

                # Mirror createIndexFromPdf-dias.py exactly:
                # keep overwriting afm on every 9-digit hit until amka is found
                for token in tokens:
                    if _AFM_PAT.match(token):
                        afm = token.lstrip('0')
                        afm_found = True
                    elif len(token) == 11 and token.isdigit():
                        amka = token.lstrip('0')
                        amka_found = True
                    if afm_found and amka_found:
                        break

                rows.append([afm, amka, i + 1])
                count += 1

                if i % 20 == 0 or i == total - 1:
                    pct = 66 + int((i + 1) / total * 34)
                    _set(yd_id, progress=min(pct, 99),
                         detail=(f'Βήμα 3/3: Σελίδα {i+1}/{total}  '
                                 f'ΑΦΜ: {afm or "—"}  ΑΜΚΑ: {amka or "—"}'))

        os.makedirs(os.path.dirname(csv_path), exist_ok=True)
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerows(rows)

        matched = sum(1 for r in rows[1:] if r[0] and r[1])
        return True, f'{matched} εγγραφές με ΑΦΜ+ΑΜΚΑ ({count} σελίδες).', matched

    except _Abort:
        raise
    except Exception as e:
        logger.exception('create_index failed')
        return False, str(e), 0


# ---------------------------------------------------------------------------
# Step 3b: Create CSV index from Excel (Strategy B)
# ---------------------------------------------------------------------------
def _convert_xls_to_xlsx(xls_path):
    """
    Convert .xls to .xlsx using LibreOffice (headless) so openpyxl can read it.
    Returns path to a temporary .xlsx file.
    """
    lo = _find_libreoffice()
    if not lo:
        raise RuntimeError('LibreOffice not found for .xls conversion')

    tmp_dir = tempfile.mkdtemp(prefix='xls2xlsx_')
    lo_profile_dir = tempfile.mkdtemp(prefix='lo_profile_conv_')
    
    # Build profile URI
    lo_profile_posix = lo_profile_dir.replace(os.sep, '/')
    if not lo_profile_posix.startswith('/'):
        lo_profile_posix = '/' + lo_profile_posix
    lo_profile_uri = f'file://{lo_profile_posix}'

    try:
        if sys.platform != 'win32':
            subprocess.run(['pkill', '-9', '-f', 'soffice'], capture_output=True)

        subprocess.run(
            [
                lo, '--headless', '--norestore', '--nofirststartwizard',
                f'-env:UserInstallation={lo_profile_uri}',
                '--convert-to', 'xlsx',
                '--outdir', tmp_dir,
                xls_path,
            ],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            check=True, timeout=1200,
            env={**os.environ, 'HOME': lo_profile_dir} if sys.platform != 'win32' else None,
        )

        out_name = os.path.splitext(os.path.basename(xls_path))[0] + '.xlsx'
        out_path = os.path.join(tmp_dir, out_name)
        if not os.path.exists(out_path):
            raise RuntimeError('LibreOffice conversion failed (no output file)')
        
        # Move to a standalone temp file
        fd, final_path = tempfile.mkstemp(suffix='.xlsx', prefix='converted_')
        os.close(fd)
        shutil.move(out_path, final_path)
        return final_path

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        shutil.rmtree(lo_profile_dir, ignore_errors=True)


def create_index_from_excel(excel_path, clean_pdf_path, csv_path, yd_id=None):
    """
    Strategy B: Read AFM+AMKA directly from the Excel file (XLS or XLSX),
    then map each employee to their PDF page range.
    
    For .xls files, they are converted to temporary .xlsx first.

    CSV format: afm;amka;page;num_pages
      page      = first PDF page of the certificate (1-based, odd)
      num_pages = pages to extract for that certificate (always 2 for XLSX)

    Returns (success, msg, matched_count).
    """
    import openpyxl

    PAGES_PER_CERT = 2
    temp_xlsx = None
    target_path = excel_path

    try:
        # If .xls, convert to .xlsx temp file
        if excel_path.lower().endswith('.xls'):
            try:
                _set(yd_id, detail='Βήμα 3/3: Μετατροπή .xls σε .xlsx για ανάγνωση δεδομένων…')
                temp_xlsx = _convert_xls_to_xlsx(excel_path)
                target_path = temp_xlsx
            except Exception as e:
                logger.warning('Could not convert .xls to .xlsx: %s', e)
                return create_index(clean_pdf_path, csv_path, yd_id)

        wb = openpyxl.load_workbook(target_path, read_only=True, data_only=True)
        records = []

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if _aborted(yd_id):
                    raise _Abort()

                afm = amka = ''
                for val in row:
                    if val is None or isinstance(val, bool):
                        continue
                    # Normalise to string — integers lose no precision
                    if isinstance(val, float):
                        int_val = int(val)
                        if float(int_val) != val:
                            continue  # has decimal part — not an ID
                        s = str(int_val)
                    elif isinstance(val, int):
                        s = str(val)
                    else:
                        s = str(val).strip()

                    # Robustly find AFM/AMKA even if mixed with text (e.g., "ΑΦΜ: 123456789")
                    # \b ensures we match whole numbers only.
                    tokens = re.findall(r'\b\d+\b', s)
                    for token in tokens:
                        if len(token) in (8, 9):
                            afm = token.lstrip('0')  # Keep last found 8 or 9-digit number
                        elif len(token) in (10, 11) and not amka:
                            amka = token.lstrip('0') # Keep first found 10 or 11-digit number

                    # Optimization: If we found both, stop scanning this row to prevent
                    # subsequent columns (like dates or amounts) from overwriting the AFM.
                    if afm and amka:
                        break

                if afm and amka:
                    records.append((afm, amka))

        wb.close()

        if not records:
            return False, 'Δεν βρέθηκαν εγγραφές ΑΦΜ+ΑΜΚΑ στο αρχείο Excel.', 0

        # Determine pages-per-certificate from actual PDF page count
        with open(clean_pdf_path, 'rb') as f:
            total_pages = len(PyPDF2.PdfReader(f).pages)

        n = len(records)
        if total_pages == n:
            PAGES_PER_CERT = 1
        elif total_pages == n * 2:
            PAGES_PER_CERT = 2
        else:
            logger.warning(
                'create_index_from_excel: PDF has %d pages, records=%d '
                '(not 1× or 2×). Falling back to PDF scan.',
                total_pages, n
            )
            _set(yd_id, detail=(
                f'Βήμα 3/3: Αναντιστοιχία σελίδων ({total_pages} PDF, {n} εγγραφές). Σάρωση PDF…'
            ))
            return create_index(clean_pdf_path, csv_path, yd_id)

        logger.info('create_index_from_excel: %d employees, %d pages/cert', n, PAGES_PER_CERT)

        # Write CSV: afm;amka;page;num_pages
        rows = [['afm', 'amka', 'page', 'num_pages']]
        for i, (afm, amka) in enumerate(records):
            page_start = i * PAGES_PER_CERT + 1  # 1-based; works for both 1 and 2 pages/cert
            rows.append([afm, amka, page_start, PAGES_PER_CERT])

            if i % 100 == 0 or i == len(records) - 1:
                pct = 66 + int((i + 1) / len(records) * 34)
                _set(yd_id, progress=min(pct, 99),
                     detail=(f'Βήμα 3/3: Εγγραφή {i+1}/{len(records)}  '
                             f'ΑΦΜ: {afm or "—"}  ΑΜΚΑ: {amka or "—"}'))

        os.makedirs(os.path.dirname(csv_path), exist_ok=True)
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerows(rows)

        return (True,
                f'{len(records)} εγγραφές (Excel-indexing, {PAGES_PER_CERT} σελίδες/βεβαίωση).',
                len(records))

    except _Abort:
        raise
    except Exception as e:
        logger.exception('create_index_from_excel failed')
        return False, str(e), 0
    finally:
        if temp_xlsx and os.path.exists(temp_xlsx):
            try: os.remove(temp_xlsx)
            except OSError: pass

# ---------------------------------------------------------------------------
# Full pipeline (background thread)
# ---------------------------------------------------------------------------
def run_pipeline(year_data_id: int, flask_app):
    def _run():
        try:
            with flask_app.app_context():
                from app import db
                from app.models import YearData

                yd = db.session.get(YearData, year_data_id)
                if not yd:
                    return

                _set(year_data_id, progress=0, detail='Προετοιμασία…', abort=False)

                data_folder = flask_app.config['DATA_FOLDER']
                year        = yd.year
                ydir        = os.path.join(data_folder, str(year))
                excel_path  = os.path.join(ydir, 'excel', yd.excel_filename)
                raw_pdf     = os.path.join(ydir, 'pdf',   f'{year}_raw.pdf')
                clean_path  = os.path.join(ydir, 'pdf',   f'{year}_clean.pdf')
                csv_path    = os.path.join(ydir, 'csv',   f'{year}.csv')

                try:
                    # ── Step 1: Excel → PDF (LibreOffice, landscape) ──────────
                    _set(year_data_id, progress=0,
                         detail='Βήμα 1/3: Εκκίνηση LibreOffice (landscape)…')
                    yd.processing_message = 'Βήμα 1/3: Μετατροπή Excel → PDF (landscape)…'
                    db.session.commit()

                    ok, msg, pages = excel_to_pdf(excel_path, raw_pdf, year_data_id)
                    if not ok:
                        raise RuntimeError(msg)

                    yd.processing_message = f'✓ Βήμα 1/3: {msg}'
                    db.session.commit()

                    # ── Step 2: Clean blank pages ─────────────────────────────
                    _set(year_data_id, progress=34,
                         detail='Βήμα 2/3: Καθαρισμός κενών σελίδων…')
                    yd.processing_message = 'Βήμα 2/3: Καθαρισμός κενών σελίδων…'
                    db.session.commit()

                    ok, msg, kept = clean_pdf(raw_pdf, clean_path, year_data_id)
                    if not ok:
                        raise RuntimeError(msg)

                    yd.processing_message = f'✓ Βήμα 2/3: {msg}'
                    db.session.commit()

                    # ── Step 3: Create CSV index ───────────────────────────────
                    _set(year_data_id, progress=67,
                         detail='Βήμα 3/3: Δημιουργία ευρετηρίου…')
                    yd.processing_message = 'Βήμα 3/3: Δημιουργία ευρετηρίου…'
                    db.session.commit()

                    ok, msg, count = create_index_from_excel(
                        excel_path, clean_path, csv_path, year_data_id
                    )
                    if not ok:
                        raise RuntimeError(msg)

                    yd.pdf_filename       = f'{year}_clean.pdf'
                    yd.csv_filename       = f'{year}.csv'
                    yd.employee_count     = count
                    yd.processing_status  = 'done'
                    yd.processing_message = f'✓ Ολοκληρώθηκε. {count} εγγραφές.'
                    yd.processed_at       = _now_athens()
                    _set(year_data_id, progress=100,
                         detail=f'✓ Ολοκληρώθηκε! {count} βεβαιώσεις έτοιμες.')

                except _Abort:
                    for path in (raw_pdf, clean_path, csv_path):
                        try:
                            if os.path.exists(path):
                                os.remove(path)
                        except OSError:
                            pass
                    yd.processing_status  = 'idle'
                    yd.processing_message = 'Ακυρώθηκε από τον χρήστη.'
                    _set(year_data_id, progress=0, detail='Ακυρώθηκε.')

                except Exception as exc:
                    logger.exception('Pipeline failed for year %s', yd.year)
                    yd.processing_status  = 'error'
                    yd.processing_message = str(exc)
                    _set(year_data_id, progress=0, detail=f'Σφάλμα: {exc}')

                finally:
                    db.session.commit()

        finally:
            release_processing_slot()

    t = threading.Thread(target=_run, daemon=True)
    with _threads_lock:
        _threads[year_data_id] = t
    t.start()
